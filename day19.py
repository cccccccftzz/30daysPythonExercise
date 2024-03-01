#File Handling

print("Catch up in Preview".center(80, "-"))
#File Handling by using open() function
f = open('reading_file_example.txt') # mode(r, a, w, x, t,b)  could be to read, write, update
print(f)

#To read the whole file
txt = f.read()
print(type(txt))

print(txt)
f.close()

#read(XX): To read the limited characters
f = open('reading_file_example.txt') 
txt = f.read(10)
print(txt)
f.close()

#readline(): To only read the first line
f = open('reading_file_example.txt') 
line = f.readline()
print(line)
f.close()

#readlines(): To only read the whole txt LINE BY LINE
f = open('reading_file_example.txt') 
lines = f.readlines()
print(lines)
f.close() #['This is a sample file to try open a file\n', 'the second line of the file\n', '\n']

#splitlines() to get all the linkes as a list
f = open('reading_file_example.txt') 
lines = f.read().splitlines()
print(type(lines)) #List
print(lines) #['This is a sample file to try open a file', 'the second line of the file', '']
f.close()

#Another way to open and close the file: with XXXXX as f:
with open('reading_file_example.txt') as f:
    lines = f.read().splitlines()
    print(type(lines)) #List
    print(lines) #['This is a sample file to try open a file', 'the second line of the file', '']

#Open file for writing and updating
with open('reading_file_example.txt', 'a') as f: #'a' means append， if the file does not exist it creates a new file
    f.write('This text has to be appended at the end.')

with open('reading_file_example_2.txt', 'w') as f: #'w' means write， if the file does not exist it creates a new file
    f.write('This text has to be appended at the end.')

#Deleting files by using os.remove
import os
os.remove('reading_file_example_2.txt')

#To use condition to print out the case if the file does not exist
if os.path.exists('reading_file_example_3.txt'):
    os.remove('reading_file_example_3.txt')
else:
    print('The file does not exist')

# Different File types
print('File with json Extension (Javascript object or Python dictionary'.center(80, "-"))
# json may use for printing the dictionary 

#Python normal dictionary 
person_dct = {
    'name' : 'Fang Ting',
    'country' : 'Singapore',
    'city' : 'Singapore',
    'skills' : ['Bodybuilding', 'programming', 'design']
}

# JSON: A string form a dictionary
person_json = '''{
    'name' : 'Fang Ting',
    'country' : 'Singapore',
    'city' : 'Singapore',
    'skills' : ['Bodybuilding', 'programming', 'design']
}'''

print('Changing JSON to dictionary'.center(80, "-"))
import json

person_json = '''{
    "name" : "Fang Ting",
    "country" : "Singapore",
    "city" : "Singapore",
    "skills" : ["Bodybuilding", "programming", "design"]
}''' #Create json string format
#Must use "" not ' ' for json module strings

person_dct = json.loads(person_json)# use json.loads() to Change JSON to dictionary
print(type(person_dct))
print(person_dct)
print(person_dct['name'])

print('Change dictionary to JSON'.center(80, "-"))
import json
person = {
    'name' : 'Fang Ting',
    'country' : 'Singapore',
    'city' : 'Singapore',
    'skills' : ['Bodybuilding', 'programming', 'design']
} #python dictionary

person_json = json.dumps(person, indent = 4) #indent could be 2, 4, 8. 
#using json.dumps() to change dictionaryto JSON
print(type(person_json))
print(person_json)

print(f'saving as JSON file'.center(80, "-"))
person = {
    'name' : 'Fang Ting',
    'country' : 'Singapore',
    'city' : 'Shanghai',
    'skills' : ['Bodybuilding', 'programming', 'design']
} #python dictionary
with open('writing_json_example', 'w', encoding = 'utf-8') as f:
    json.dump(person, f, ensure_ascii=False, indent = 4)

''' encoding='utf-8' parameter specifies that the file should be opened with UTF-8 encoding. 
UTF-8 is a widely used character encoding that can represent most of the characters in the Unicode character set
'''

print('file with csv Entension'.center(80, "-"))
#csv used to store tabular data i.e. spreadsheet or database
import csv
with open ('csv_sample.csv') as f:
    csv_reader = csv.reader(f, delimiter = ',') #csv.reader is same as 'w' use
    line_count = 0
    for row in csv_reader:
        if line_count == 0:
            print(f'Column names are: {",".join(row)}')
            #row is the first title row
            line_count += 1
        else: 
            print(
                f'\t{row[0]} is a teacher. He lives in {row[1]}, {row[2]}.'
            )    
            #row[index], index refer to column number, start from column 0 ~2
            line_count += 1
        print(f'Number of lines: {line_count}')

print('file with xlsx Entension'.center(80, "-"))
# import xlrd # To read the excel files we need to install xlrd package
# excel_book = xlrd.open_workbook('ST21_Drawing List 2.xlsx')
# print(excel_book.nsheets)
# print(excel_book.sheet_names)
              
from openpyxl import load_workbook

# workbook = load_workbook(filename='ST21_Drawing List 2.xlsx')
# wb = workbook.active
# print(sheet.nsheets)
# print(sheet.sheetnames)
# print(wb.get_sheet_names())

# Now you can read or write to the workbook
                            
print('file with xml extension'.center(80, "-"))
import xml.etree.ElementTree as ET
tree = ET.parse('xml_example.xml')
root = tree.getroot()
print('Root tag:', root.tag)
print('Attribute:', root.attrib)
for child in root:
    print('field: ', child.tag)

print('Exercise Level 1'.center(80, "-"))
print('Exercise 1'.center(80, "-"))

print(f'a) Read obama_speech.txt file and count number of lines and words')
import re
from collections import Counter

def count_rows_and_words_in_text(file_path):
    f = open(file_path) 
    lines = f.read().splitlines()
    print(f'The total rows in the file is {len(lines)}') #Count the total rows in the txt
    word_counts = 0
    for sentence in lines:
        regex_pattern = r'\b\w+\b' #\b to find a word boundary, \w to find one word character,+ means one or more times
        words = re.findall(regex_pattern, sentence)
        word_counts += len(words)
    print(f'The total words number is {word_counts}')
    f.close()

count_rows_and_words_in_text('./data/obama_speech.txt')

print(f'b) Read michelle_obama_speech.txt file and count number of lines and words')
count_rows_and_words_in_text('./data/michelle_obama_speech.txt')


print(f'c) Read donald_speech.txt file and count number of lines and words')
count_rows_and_words_in_text('./data/donald_speech.txt')


print(f'd) Read melina_trump_speech.txt file and count number of lines and words')
count_rows_and_words_in_text('./data/melina_trump_speech.txt')

print('Exercise 2'.center(80, "-"))
def find_top_languages(json_file_path, num): # num refer to how many numbers of the top
    with open(json_file_path, 'r', encoding = 'utf-8') as file:
        countries_data = json.load(file) #By using json.load() to load the file to dictionary
    # Argument : It takes file object as a parameter. Return : It return json object.
    #If it takes the list as input, it return a list also
    # print(type(countries_data))
    all_language = [language for country in countries_data for language in country['languages']]
    language_count = Counter(all_language)
    top_languages = language_count.most_common(num)
    reverse_top_languages = [(counts, language) for language, counts in top_languages]
    return reverse_top_languages

file_path = './data/countries_data.json'
top_num_languages = find_top_languages(file_path, 10)
print(top_num_languages)

file_path = './data/countries_data.json'
top_num_languages = find_top_languages(file_path, 3)
print(top_num_languages)


print('Exercise 3'.center(80, "-"))
print('Read the countries_data.json data file in data directory, create a function that creates a list of the ten most populated countries')
from collections import defaultdict
def most_populated_countries(json_file_path, num):
    with open(json_file_path, 'r', encoding = 'utf-8') as file:
        countries_data = json.load(file)
    country_population = {}
    country_population = [{'country':country['name'], 'population': country['population']} for country in countries_data]
    # for country in countries_data:
    #     country_population['country'] = country['name']
    #     country_population['population'] = country['population']
    sorted_country_population = sorted(country_population, key = lambda x: x["population"], reverse = True)
    print(sorted_country_population[:num])

most_populated_countries('./data/countries_data.json', 10)
most_populated_countries('./data/countries_data.json', 3)

'''
sample result
[
{'country': 'China', 'population': 1377422166},
{'country': 'India', 'population': 1295210000},
{'country': 'United States of America', 'population': 323947000},
{'country': 'Indonesia', 'population': 258705000},
{'country': 'Brazil', 'population': 206135893},
{'country': 'Pakistan', 'population': 194125062},
{'country': 'Nigeria', 'population': 186988000},
{'country': 'Bangladesh', 'population': 161006790},
{'country': 'Russian Federation', 'population': 146599183},
{'country': 'Japan', 'population': 126960000}
]
'''

print('Exercise Level 2'.center(80, "-"))
print('Exercise 4'.center(80, "-"))
print(f'Extract all incoming email addresses as a list from the email_exchange_big.txt file.')
with open('./data/email_exchanges_big.txt') as file:
    txt = file.read()
    regex_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}' #\.: matches the dot separating the domain name from the top-level domain (TLD)
    #No \ for first few character options: regex_pattern = r'[^A-Za-z]+' 
    email_address = re.findall(regex_pattern, txt)
    # print(email_address)

print('Exercise 5'.center(80, "-"))
print('''Find the most common words in the English language. 
Call the name of your function find_most_common_words, 
it will take two parameters - a string or a file and a positive integer, 
indicating the number of words. Your function will return an array of tuples in descending order. ''')

import os
 
 #Def a function to tell the input is a path or string directly
def is_file_path(input):
    return os.path.isfile(input)

#The main function to find the top most xx word
def find_most_common_words(file_path_or_str, num):
    if is_file_path(file_path_or_str):
        with open(file_path_or_str, 'r', encoding = 'utf-8') as file:
            txt = file.read()
    else:
        txt = file_path_or_str

    regex_pattern = r'\b\w+\b'
    words_lst = re.findall(regex_pattern, txt)
    word_count = Counter(words_lst)
    most_common_words = word_count.most_common(num)
    reversed_most_common_words = [(count, word) for word, count in most_common_words]
    print(reversed_most_common_words)

find_most_common_words('./files/example.txt', 10)
find_most_common_words('./files/example.txt', 5)

print('Exercise 6'.center(80, "-"))
# find_most_common_words('.\data\obama_speech.txt', 10)
''' The file path here got syntax error bc the backslash 
The error you're encountering is due to the backslashes in the file path being interpreted as escape characters in the string. 
To resolve this, you can use raw string literals (prefix the string with r) or double backslashes. 
see the solutions below
'''

find_most_common_words(r'.\data\obama_speech.txt', 10) #Solution 1: start with r'xxx'
find_most_common_words('.\\data\\michelle_obama_speech.txt', 10) #Solution 2: use double \\
find_most_common_words(r'.\data\donald_speech.txt', 10)
find_most_common_words(r'.\data\melina_trump_speech.txt', 10)

print('Exercise 7'.center(80, "-"))
print('''Write a python application that checks similarity between two texts. 
It takes a file or a string as a parameter and it will evaluate the similarity of the two texts. 
You may need a couple of functions, function to clean the text(clean_text), function to remove support words(remove_support_words) 
and finally to check the similarity(check_text_similarity). List of stop words are in the data directory
''')

#To def a function to judege whether the input is a file path or string
def is_file_path(input):
    return os.path.isfile(input)

#To def a function to clean the text & return a list of the words included in the txt
def clean_text(txt):
    regex_pattern = r'\b\w+\b'
    word_list = re.findall(regex_pattern, txt)
    return word_list

    '''
    Shall not convert to set in this stage to make the following function more useful     
    word_set = set(word_lst) 
    '''

#To def a function to remove the stop words: To take one set input then return a set without stop words
from data.stop_words import stop_words

def remove_support_words(word_list): 
    set_stop_words = set(stop_words)
    word_set = set(word_list)
    return word_set - set_stop_words 
#Can directly to use the '-' to minus those stop words

#My own mthod
# def remove_support_words(word_set):
#     word_list = list(word_set)
#     for word in word_list:
#         if word in stop_words:
#             # word_list = word_list.remove(word) #Wrong, bc lst.remove() will edit the list in place
#             word_list.remove(word)
#     word_set_no_support_word = set(word_list)
#     return word_set_no_support_word

#To def a function to check the similarity between two txts
def check_text_similarity(text_or_path_one, text_or_path_two):
    
    #To tell the input text or file path for two arguments
    if is_file_path(text_or_path_one):
        with open(text_or_path_one, 'r', encoding = 'utf-8') as file_one:
            txt_one = file_one.read()
    else:
        txt_one = text_or_path_one

    if is_file_path(text_or_path_two):
        with open(text_or_path_two, 'r', encoding = 'utf-8') as file_two:
            txt_two = file_two.read()
    else:
        txt_two = text_or_path_two

    #To clear the txt and return sets of words for two txts
    word_set_one = clean_text(txt_one)
    word_set_two = clean_text(txt_two)
    
    #To remove the support words in two sets of words
    word_set_one_no_support_words = remove_support_words(word_set_one)
    word_set_two_no_support_words = remove_support_words(word_set_two)

    #Similarity = commonword/total word
    common_word = word_set_one_no_support_words.intersection(word_set_two_no_support_words)
    count_common_word = len(common_word)

    superset_one_and_two = word_set_one_no_support_words.union(word_set_two_no_support_words)
    count_superset_word = len(superset_one_and_two)

    count_word_txt_one = len(word_set_one_no_support_words)
    count_word_txt_two = len(word_set_two_no_support_words)

    similarity = count_common_word / count_superset_word
    print(f'The similarity between two text is: {similarity}')

#To run with the two text paths or just give two strings
text_or_path_one=r'.\data\michelle_obama_speech.txt'
text_or_path_two=r'.\data\melina_trump_speech.txt'
check_text_similarity(text_or_path_one, text_or_path_two)


print('Exercise 8'.center(80, "-"))
print(f'Find the 10 most repeated words in the romeo_and_juliet.txt')
def most_repeated_words(file_path_or_str, num):
    
    #To tell the input text or file path
    if is_file_path(file_path_or_str):
        with open(file_path_or_str, 'r', encoding = 'utf-8') as file:
            txt = file.read()
    else:
        txt= file_path_or_str

    #To clear the txt and return the list of words in the text
    word_list = clean_text(txt) #Output is the total words list
    word_list_lower = [word.lower() for word in word_list]

    #To remove the support words in the list regardless of the captial by changing all the words in the list to lower captial
    word_list_lower = [word for word in word_list_lower if word not in stop_words] #To replace with a list comprehension to avoid a return statmenet

    '''
    Wrong method, to give a reutrn value inside a loop in a function
    The issue in your code is that the return statement inside the loop is causing the function to exit prematurely.
    for word in word_list:
        if word in stop_words:
            word_list.remove(word)
            return word_list
    '''
    word_count = Counter(word_list_lower)
    most_repeated_words = word_count.most_common(num)
    print(most_repeated_words)

most_repeated_words(r'.\data\romeo_and_juliet.txt', 10)
    
print('Exercise 9'.center(80, "-"))
print('''Read the hacker news csv file and find out the following text''')

print(f'a) Count the number of lines containing python or Python ')
with open(r'.\data\hacker_news.csv') as file:
    csv_reader = csv.reader(file, delimiter = ',')
    
    line_count = 0
    count_python = 0

    for row in csv_reader:
        line_count += 1
        for element in row:
            regex_pattern = r'\b[Pp]ython\b'
            matches = re.search(regex_pattern, element)
            #Or to use the re.search but with flag
            # matches = re.search(regex_pattern, element, flags=re.IGNORECASE)  # Ignore case
            if matches is not None:
                count_python += 1
                # print(f'{line_count = }') #Can print out the corresponding line number
                # print(row) #Can print out the corresponding row content
                break 
                #if already find the key word in one element, then end the loop to the next row

    print(count_python)

print('b) Count the number lines containing JavaScript, javascript or Javascript')
with open(r'.\data\hacker_news.csv') as file:
    csv_reader = csv.reader(file, delimiter = ',')
    
    line_count = 0
    count_javascript = 0

    for row in csv_reader:
        line_count += 1
        for element in row:
            regex_pattern = r'\b[Jj]ava[Ss]cript\b'
            matches = re.search(regex_pattern, element)
            #Or to use the re.search but with flag
            # matches = re.search(regex_pattern, element, flags=re.IGNORECASE)  # Ignore case
            if matches is not None:
                count_javascript += 1
                # print(f'{line_count = }') #Can print out the corresponding line number
                # print(row) #Can print out the corresponding row content
                break 
                
    print(count_javascript)

print('c) Count the number lines containing Java and not JavaScript')
with open(r'.\data\hacker_news.csv') as file:
    csv_reader = csv.reader(file, delimiter = ',')
    
    line_count = 0
    count_java = 0

    for row in csv_reader:
        line_count += 1
        for element in row:
            regex_pattern = r'\b[Jj]ava\b'
            matches = re.search(regex_pattern, element)
            #Or to use the re.search but with flag
            # matches = re.search(regex_pattern, element, flags=re.IGNORECASE)  # Ignore case
            if matches is not None:
                count_java += 1
                print(f'{line_count = }') #Can print out the corresponding line number
                print(row) #Can print out the corresponding row content
                break 
               
    print(count_java)